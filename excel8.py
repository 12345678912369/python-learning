import openpyxl

#创建一个新的Excel工作簿
wb = openpyxl.Workbook()
#选择默认的工作表
ws = wb.active  
ws.title = "学生成绩"  #给工作表命名

#写入表头
ws['A1'] = "姓名"
ws['B1'] = "成绩"
ws['C1'] = "是否及格"
#写入数据
students = [("张三", 85),("李四", 78), ("王五", 92), ("赵六", 55)]
for i, student in enumerate(students):#行开始写数据
 ws["A"+str(i+2)] = student[0]
 ws["B"+str(i+2)] = student[1]
 ws["C"+str(i+2)] = "及格" if student[1] >= 60 else "不及格"
#保存工作簿
wb.save("学生成绩.xlsx")
print("数据已写入学生成绩.xlsx文件")