'''requests 是抓网页内容，BeautifulSoup 是把抓回来的HTML解析成可以用的数据，两个配合使用才是完整的爬虫。
就像：
requests = 把整本书拿回来
BeautifulSoup = 找到书里你要的那一页'''



'''from bs4 import BeautifulSoup
import requests
response = requests.get(                #抓取网页
    "https://books.toscrape.com",
    proxies={"http": "http://127.0.0.1:7897", "https": "http://127.0.0.1:7897"}
)
#解析HTML
soup = BeautifulSoup(response.text, "html.parser")
#找到所有的书名
books = soup.find_all("article",class_="product_pod")
for book in books:
    title = book.find("h3").find("a")["title"]  #书名在h3标签的a标签的title属性里
    price = book.find("p", class_="price_color").text  #价格在p标签的price_color类里
    print(f"书名: {title}, 价格: {price}")#print(title}-{price})'''


'''requests.get() 发出请求后，服务器返回的所有东西都存在 response 里，包括：
response.status_code — 状态码
response.text — 网页内容
response.json() — JSON数据
叫 response 是因为它存的是服务器的响应结果，你也可以叫 r、res、result，效果一样


BeautifulSoup这个库名字本身就来自一个典故——把杂乱的HTML"煮成一锅汤"，方便你从里面捞东西。
所以变量习惯叫 soup，表示解析好的HTML对象，之后用它来找标签：
soup.find() — 找一个
soup.find_all() — 找所有


response是原始网页，soup是解析后可以操作的网页。 先抓再解析，固定搭配'''


#结合openpyxl把爬到的书名和价格写到Excel里
from bs4 import BeautifulSoup
import requests
import openpyxl
response = requests.get(
    "https://books.toscrape.com",
    proxies={"http": "http://127.0.0.1:7897", "https": "http://127.0.0.1:7897"}
)
soup = BeautifulSoup(response.text, "html.parser")
books = soup.find_all("article",class_="product_pod")
#创建Excel工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "书籍信息"
#写入表头
ws['A1'] = "书名"
ws['B1'] = "价格"
#写入数据
for i, book in enumerate(books):
    title = book.find("h3").find("a")["title"]
    price = book.find("p", class_="price_color").text
    ws["A"+str(i+2)] = title
    ws["B"+str(i+2)] = price
#保存工作簿
wb.save("书籍信息.xlsx")
print("书籍信息已写入书籍信息.xlsx文件，共"+str(len(books))+"本)")