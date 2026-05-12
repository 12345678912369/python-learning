import requests
from bs4 import BeautifulSoup
import openpyxl
import os

# 代理地址（如果不需要代理，设为 None）
PROXY = None   # 示例: {"http": "http://127.0.0.1:7897", "https": "http://127.0.0.1:7897"}
EXCEL_FILE = "books.xlsx"

def scrape_books():
    """爬取书籍数据"""
    print("正在采集书籍数据...")

    try:
        if PROXY:
            response = requests.get("https://books.toscrape.com", proxies=PROXY, timeout=10)
        else:
            response = requests.get("https://books.toscrape.com", timeout=10)
    except requests.RequestException as e:
        print(f"网络请求失败：{e}")
        return []

    soup = BeautifulSoup(response.text, "html.parser")
    items = soup.find_all("article", class_="product_pod")

    books = []
    for item in items:
        title_tag = item.find("h3")
        title = title_tag.find("a")["title"] if title_tag else "未知书名"

        price_tag = item.find("p", class_="price_color")
        price = price_tag.text.strip() if price_tag else "未知价格"

        books.append({"title": title, "price": price})

    print(f"采集完成！共 {len(books)} 本书")
    save_to_excel(books)
    return books

def save_to_excel(books):
    """保存数据到 Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "书籍列表"
    ws.append(["序号", "书名", "价格"])

    for i, book in enumerate(books, 1):
        ws.append([i, book["title"], book["price"]])

    wb.save(EXCEL_FILE)
    print(f"数据已保存至 {EXCEL_FILE}")

def view_data():
    """查看已保存的数据"""
    if not os.path.exists(EXCEL_FILE):
        print("还没有数据，请先选择 1 采集！")
        return

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    print("\n=== 已采集的书籍 ===")

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        print("暂无数据。")
        wb.close()
        return

    for row in rows:
        print(f"  {row[0]}. {row[1]} - {row[2]}")

    wb.close()
    print(f"\n共 {len(rows)} 条记录")

def show_menu():
    """主菜单"""
    actions = {"1": scrape_books, "2": view_data}

    while True:
        print("\n=== 书籍数据采集工具 ===")
        print("1. 数据采集")
        print("2. 数据查看")
        print("3. 退出")

        choice = input("请选择操作：").strip()

        if choice == "3":
            print("退出程序，再见！")
            break

        action = actions.get(choice)
        if action:
            action()
        else:
            print("输入无效，请输入 1-3。")

if __name__ == "__main__":
    show_menu()
